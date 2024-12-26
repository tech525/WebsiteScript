from flask import Flask, request, send_file
import os
import pandas as pd
from openpyxl import Workbook
from datetime import timedelta
import yfinance as yf
from openpyxl.styles import PatternFill

app = Flask(__name__)

# Function to fetch dividend data
def get_dividend_data(stock, num_stocks):
    try:
        ticker = yf.Ticker(stock)
        dividends = ticker.dividends

        if dividends.empty:
            return [stock, "No dividend data available", "-", "-", "-", "-"]

        last_dividend_date = dividends.index[-1] if not dividends.index.empty else None
        last_dividend_amount = dividends.iloc[-1] if len(dividends) > 0 else None

        if not last_dividend_date:
            return [stock, "Invalid Dividend Date", "-", "-", "-", "-"]

        last_dividend_date_str = last_dividend_date.strftime('%d %b %Y')
        likely_credit_date = calculate_likely_credit_date(last_dividend_date_str)
        total_dividend = num_stocks * last_dividend_amount if last_dividend_amount else "-"

        return [stock, last_dividend_date_str, last_dividend_amount, likely_credit_date, num_stocks, total_dividend]

    except Exception as e:
        return [stock, "Error", "-", "-", "-", "-"]

# Function to calculate likely credit date
def calculate_likely_credit_date(last_dividend_date):
    try:
        last_dividend_date = pd.to_datetime(last_dividend_date, format='%d %b %Y')
        working_days_count = 0
        current_date = last_dividend_date

        while working_days_count < 25:
            current_date += timedelta(days=1)
            if current_date.weekday() < 5:  # Monday to Friday
                working_days_count += 1

        return current_date.strftime('%d %b %Y')
    except Exception as e:
        return "-"

# Function to save data to Excel
def save_to_excel(dividend_data, output_file):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dividend Tracker"

    headers = [
        "Stock Name",
        "Dividend Date",
        "Dividend Amount (INR)",
        "Likely Credit Date",
        "Number of Stocks",
        "Total Dividend Amount (INR)"
    ]
    sheet.append(headers)

    today = pd.to_datetime("today").date()

    for row_index, row in enumerate(dividend_data, start=2):
        sheet.append(row)
        likely_credit_date = row[3]
        if likely_credit_date != "-" and pd.to_datetime(likely_credit_date, format='%d %b %Y').date() > today:
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            sheet.cell(row=row_index, column=4).fill = fill

    workbook.save(output_file)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "No file uploaded", 400

    file = request.files['file']
    if file.filename == '':
        return "No selected file", 400

    input_path = "input.xlsx"
    output_path = "output.xlsx"
    file.save(input_path)

    try:
        stock_data = pd.read_excel(input_path, usecols=["Stock Name", "Number of Stocks"])
        stock_data = stock_data.dropna(subset=["Stock Name", "Number of Stocks"])
        stock_data["Number of Stocks"] = pd.to_numeric(stock_data["Number of Stocks"], errors="coerce").fillna(0).astype(int)
        stock_data = stock_data[stock_data["Number of Stocks"] > 0]

        stocks = stock_data["Stock Name"].tolist()
        num_stocks = stock_data["Number of Stocks"].tolist()

        dividend_data = []
        for stock, num in zip(stocks, num_stocks):
            data = get_dividend_data(stock, num)
            dividend_data.append(data)

        save_to_excel(dividend_data, output_path)
        return send_file(output_path, as_attachment=True)

    except Exception as e:
        return f"Error processing file: {e}", 500

    finally:
        if os.path.exists(input_path):
            os.remove(input_path)
        if os.path.exists(output_path):
            os.remove(output_path)

@app.route('/')
def index():
    return '''
    <!doctype html>
    <title>Upload File</title>
    <h1>Upload your Excel file to process dividends</h1>
    <form method=post enctype=multipart/form-data action="/upload">
      <input type=file name=file>
      <input type=submit value=Upload>
    </form>
    '''

if __name__ == '__main__':
    import os
    port = int(os.environ.get("PORT", 5000))  # Use Render's PORT environment variable
    app.run(host='0.0.0.0', port=port, debug=True)
